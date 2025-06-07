import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
from rapidfuzz import fuzz
import re
from datetime import datetime, timedelta
import unidecode
import os

def limpar_nome(nome):
    nome = unidecode.unidecode(nome.lower())
    for prefixo in ['atacadão', 'atacadao', 'atc', 'atacado', 'atc.', 'atacadão.', 'atacadao.', 'atacado.']:
        nome = re.sub(r'\b' + prefixo + r'\b', '', nome)
    nome = re.sub(r'[^a-zA-Z0-9 ]', '', nome)
    nome = re.sub(r'\s+', ' ', nome)
    return nome.strip()

def gerar_relatorio(caminho_excel, caminho_txt):
    NOMES_PROMOTOR = ['Rodrigo', 'P3 Trade Supervisão']
    data_inicio = datetime.strptime('26/05/2025', '%d/%m/%Y')
    dias_semana_original = ['segunda', 'terça', 'quarta', 'quinta', 'sexta', 'sábado']
    dias_semana_normalizada = [unidecode.unidecode(d).lower() for d in dias_semana_original]
    datas_semana = [(data_inicio + timedelta(days=i)).strftime('%d/%m/%Y') for i in range(6)]

    apelidos_lojas = {
        'anchieta': ['anchieta', 'atc anchieta', 'atacadão anchieta'],
        'aricanduva': ['aricanduva', 'atc aricanduva', 'atacadão aricanduva'],
        'barueri ii': ['barueri 2', 'atc barueri', 'atacadão barueri 2', 'atacadao barueri 2', 'atacadão barueri', 'atacadao barueri'],
        'carapicuiba': ['carapicuiba', 'atc carapicuiba', 'atacadão carapicuiba'],
        'cotia centro': ['cotia centro', 'atc cotia centro'],
        'cotia ii': ['cotia 2', 'atc cotia 2', 'atc cotia2', 'cotia2 atc', 'cotia2', 'cotia ii'],
        'diadema (ex maxxi) (canhema)': ['diadema', 'canhema', 'maxxi', 'diadema maxxi', 'diadema canhema'],
        'diadema cupecê': ['cupecê', 'cupece', 'diadema cupecê'],
        'diadema serraria': ['serraria', 'diadema serraria'],
        'ferraz vasconcelos': ['ferraz', 'ferraz vasconcelos'],
        'gabiroba carapicuiba': ['gabiroba', 'atc gabiroba', 'gabiroba carapicuiba'],
        'guarulhos aeroporto': ['guarulhos aeroporto', 'aeroporto'],
        'guarulhos bonsucesso': ['bonsucesso', 'guarulhos bonsucesso'],
        'guarulhos centro': ['guarulhos centro', 'centro guarulhos'],
        'guarulhos dutra': ['guarulhos dutra', 'dutra'],
        'inajar de souza': ['inajar', 'inajar de souza'],
        'itapevi': ['itapevi'],
        'itaquera': ['itaquera'],
        'jandira': ['jandira'],
        'maua': ['maua', 'mauá'],
        'maua joao ramalho': ['joão ramalho', 'joao ramalho', 'maua joao ramalho'],
        'mirandopolis': ['mirandopolis'],
        'mogi das cruzes': ['mogi das cruzes', 'mogi'],
        'osasco': ['osasco'],
        'osasco yolanda': ['yolanda', 'atc yolanda'],
        'penha': ['penha'],
        'pirituba': ['pirituba'],
        'poa': ['poa', 'poá'],
        'santo andre (ex maxxi) capuava av do estado': ['capuava', 'maxxi', 'capuava av do estado'],
        'santo andre centro - queiroz dos santos': ['queiroz', 'queiroz dos santos', 'santo andre centro'],
        'santo andre ii': ['santo andre ii'],
        'sao bernardo do campo': ['são bernardo', 'sao bernardo', 'sbc'],
        'sao caetano do sul': ['sao caetano', 'são caetano'],
        'sao miguel - marechal tito': ['são miguel', 'marechal tito'],
        'sbc demarchi': ['demarchi', 'atc demarchi', 'sbc demarchi'],
        'sbc marechal tito': ['marechal sbc', 'sbc marechal'],
        'suzano': ['suzano'],
        'taipas': ['taipas'],
        'tambore': ['tambore', 'tamboré'],
        'vila jacui': ['vila jacui', 'jacui'],
        'vila lobos': ['vila lobos'],
        'tatuape (pq. sevilha)': ['tatuape', 'tatuapé', 'pq sevilha', 'atacadão tatuape', 'atacadao tatuape', 'atacadão tatuapé', 'atacadao tatuapé'],
    }

    df = pd.read_excel(caminho_excel)
    colunas_originais = list(df.columns)
    colunas_normalizadas = [unidecode.unidecode(col).strip().lower() for col in colunas_originais]
    mapa_colunas = dict(zip(colunas_normalizadas, colunas_originais))

    with open(caminho_txt, 'r', encoding='utf-8') as f:
        linhas = f.readlines()

    msgs_por_data = {d: [] for d in datas_semana}
    padrao = r"\[(\d{2}:\d{2}), (\d{2}/\d{2}/\d{4})\] (.+?): (.+)"
    for linha in linhas:
        match = re.match(padrao, linha.strip())
        if match:
            hora, data_msg, nome, conteudo = match.groups()
            nome_normalizado = nome.lower().strip()
            if any(n.lower() in nome_normalizado for n in NOMES_PROMOTOR):
                if data_msg in msgs_por_data:
                    msgs_por_data[data_msg].append(conteudo.lower())

    for idx, row in df.iterrows():
        loja_oficial = row[mapa_colunas['loja']].lower()
        loja_limpa = limpar_nome(loja_oficial)
        palavras_loja = set(loja_limpa.split())
        aliases = apelidos_lojas.get(loja_limpa, [])
        for i, dia_norm in enumerate(dias_semana_normalizada):
            data = datas_semana[i]
            encontrou = False
            for msg in msgs_por_data[data]:
                msg_limpa = limpar_nome(msg)
                score = fuzz.ratio(loja_limpa, msg_limpa)
                palavras_msg = set(msg_limpa.split())
                if palavras_loja <= palavras_msg or score > 95:
                    encontrou = True
                    break
                for alias in aliases:
                    alias_limpo = limpar_nome(alias)
                    score_alias = fuzz.ratio(alias_limpo, msg_limpa)
                    if alias_limpo in msg_limpa or score_alias > 90:
                        encontrou = True
                        break
                if encontrou:
                    break
            col_original = mapa_colunas[dia_norm]
            if encontrou:
                df.at[idx, col_original] = 'X'
            else:
                df.at[idx, col_original] = None

    df = df[[col for col in colunas_originais]]

    base_nome = os.path.splitext(os.path.basename(caminho_txt))[0]
    caminho_saida = f"relatorio_{base_nome}.xlsx"
    df.to_excel(caminho_saida, index=False)

    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Border, Side

    wb = load_workbook(caminho_saida)
    ws = wb.active

    verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    laranja = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    dias_semana = ['segunda', 'terça', 'quarta', 'quinta', 'sexta', 'sábado']
    col_dias = []
    for col in ws.iter_cols(1, ws.max_column, 1, 1):
        cabec = col[0].value
        if cabec is not None and cabec.lower() in dias_semana:
            col_dias.append(col[0].column)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for col_idx in col_dias:
            cell = row[col_idx-1]
            if cell.value == 'X':
                cell.fill = verde
            else:
                cell.fill = laranja

    borda_fina = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000"),
    )
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = borda_fina

    lojas_sem_atendimento = []
    col_loja = None
    colunas_dias = []
    for col in ws.iter_cols(1, ws.max_column, 1, 1):
        cabec = str(col[0].value).strip().lower()
        if cabec == "loja":
            col_loja = col[0].column
        if cabec in ['segunda', 'terça', 'quarta', 'quinta', 'sexta', 'sábado']:
            colunas_dias.append(col[0].column)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        tem_atendimento = False
        for col_idx in colunas_dias:
            if row[col_idx-1].value == 'X':
                tem_atendimento = True
                break
        if not tem_atendimento and col_loja is not None:
            loja_nome = row[col_loja-1].value
            lojas_sem_atendimento.append(loja_nome)

    if lojas_sem_atendimento:
        aviso = "Atenção: as lojas abaixo não tiveram registros de atendimento nesta semana:\n" + \
                ", ".join(str(loja) for loja in lojas_sem_atendimento)
        ws.insert_rows(1)
        ws['A1'] = aviso
        aviso_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        ws['A1'].fill = aviso_fill
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)

    wb.save(caminho_saida)
    return caminho_saida

def selecionar_arquivo(entry_widget):
    filename = filedialog.askopenfilename()
    if filename:
        entry_widget.delete(0, tk.END)
        entry_widget.insert(0, filename)

def rodar():
    txt = txt_entry.get()
    excel = excel_entry.get()
    if not txt or not excel:
        messagebox.showerror("Erro", "Selecione ambos os arquivos.")
        return
    try:
        status_label.config(text="Gerando relatório...", fg="blue")
        caminho = gerar_relatorio(excel, txt)
        status_label.config(text=f"Relatório gerado: {caminho}", fg="green")
        messagebox.showinfo("Sucesso", f"Relatório gerado: {caminho}")
    except Exception as e:
        status_label.config(text=f"Erro: {e}", fg="red")
        messagebox.showerror("Erro ao gerar relatório", str(e))

root = tk.Tk()
root.title("Geração de Relatório de Visitas Eco Fresh")

tk.Label(root, text="Arquivo TXT do WhatsApp:").grid(row=0, column=0, sticky='e')
txt_entry = tk.Entry(root, width=50)
txt_entry.grid(row=0, column=1)
tk.Button(root, text="Browse", command=lambda: selecionar_arquivo(txt_entry)).grid(row=0, column=2)

tk.Label(root, text="Planilha Base Excel:").grid(row=1, column=0, sticky='e')
excel_entry = tk.Entry(root, width=50)
excel_entry.grid(row=1, column=1)
tk.Button(root, text="Browse", command=lambda: selecionar_arquivo(excel_entry)).grid(row=1, column=2)

tk.Button(root, text="Gerar Relatório", command=rodar).grid(row=2, column=1, pady=10)
status_label = tk.Label(root, text="", fg="black")
status_label.grid(row=3, column=0, columnspan=3)

root.mainloop()
